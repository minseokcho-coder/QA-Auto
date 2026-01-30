"""TC 케이스 엑셀 파일 생성 스크립트 (새 양식)"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# TC 케이스 데이터 정의 (이미지 참고하여 작성)
test_cases = [
    # 토스/카카오 진입 테스트
    {
        "No": 1,
        "preconditions": "1. 토스 경로로 진입",
        "title": "토스로 진입하는 경우 동작 확인",
        "steps_actions": "1. 토스 진입 후 화면 이탈 > 카카오링크로 재진입하는 경우 동작 확인\n2. 카카오 진입 후 화면 이탈 > 토스로 재진입하는 경우 동작 확인",
        "steps_result": "이전 레거시 동작 그대로 유지(진입 여부만 확인하기)",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 2,
        "preconditions": "1. 토스 경로로 진입",
        "title": "통합 진입 후 화면 이탈하고 다른 경로로 진입하는 경우 동작 확인",
        "steps_actions": "1. 토스 진입 후 화면 이탈하고 다른 경로로 진입하는 경우 동작 확인",
        "steps_result": "1. 토스로 진입(최초 진입한 걸 기준으로 진행)\n2. 카카오로 진입(최초 진입한 걸 기준으로 진행)",
        "priority": "Low",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 3,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입",
        "title": "세이브택스 환급진입 동작 확인",
        "steps_actions": "카카오로 계속하기 선택 시 동작 확인",
        "steps_result": "1. 신규 유저인 경우, 카카오 계정으로 간편 가입 > 동의화면 노출\n2. 기존 유저인 경우.",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 4,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입",
        "title": "카카오계정으로 간편가입 동의 시 동작 확인",
        "steps_actions": "1. 동의하고 계속하기 선택 시 동작 확인\n2. 취소 선택 시 동작 확인",
        "steps_result": "1. 홈택스 인증 화면으로 이동\n2. 동의 팝업 닫히며 이전화면으로 이동(이전 레거시 그대로 진행)",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 5,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입",
        "title": "법인 사업자인 경우 동작 확인",
        "steps_actions": "혹시 법인사업자이신가요? 선택 시 동작 확인",
        "steps_result": "세이브택스 환급 법인 페이지로 이동",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 6,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입",
        "title": "홈택스 간편 인증 동작 확인",
        "steps_actions": "1. 간편인증 수단 동작 확인\n2. 간편인증 정보 입력 동작 확인\n3. 인증 요청 동작 확인",
        "steps_result": "이전 레거시 동작 그대로 유지(수정 필요한 부분이 있다면 버그 등록 후 논의 필요)",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 7,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입",
        "title": "홈택스 약관 동의 동작 확인",
        "steps_actions": "1. 필수 동의 모두 선택하는 경우 동작 확인\n2. 필수의 모두 미선택하는 경우 동작 확인\n3. 필수의 일부 선택하는 경우 동작 확인",
        "steps_result": "1. 하단 버튼 활성화되며, 선택 시 다음 페이지(=인증 요청 메세지)로 이동\n2. 하단 버튼 비활성화\n3. 하단 버튼 비활성화",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 8,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입",
        "title": "홈택스 인증 요청 메시지 동작 확인",
        "steps_actions": "1. 인증 요청 메시지 동작 확인\n2. 인증 완료 후 클릭 버튼 선택 시 동작 확인",
        "steps_result": "이전 레거시 동작 그대로 유지(수정 필요한 부분이 있다면 버그 등록 후 논의 필요)",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 9,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입\n3. 홈택스 인증 완료한 경우",
        "title": "통합 스크래핑 브릿지 동작 확인",
        "steps_actions": "1. 화면 유지한 상태에서 스크래핑 완료된 경우 동작 확인\n2. 화면 이탈한 상태에서 스크래핑 완료된 경우 동작 확인",
        "steps_result": "1. 예상 환급액 확인 페이지로 이동\n2. 카카오 알림톡으로 완료 안내 발송",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 10,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입\n3. 홈택스 인증 완료한 경우\n4. 스크래핑 실패한 경우",
        "title": "스크래핑 실패한 경우 동작 확인",
        "steps_actions": "1. 환급액 조회 지연되는 경우 확인\n2. 홈택스 정보 불러오기에 실패하는 경우 동작 확인\n3. 홈택스 점검 시간인 경우 동작 확인\n4. 서비스 점검 중으로 환급액 계산이 지연되는 경우 확인",
        "steps_result": "1. 환급액 조회가 지연되고 있어요 노출\n2. 홈택스 정보 불러오기에 실패했어요 노출\n3. 지금은 홈택스 점검시간이에요(00시~06시)\n4. 서비스 점검중으로 환급액 계산이 지연되고 있어요.",
        "priority": "Medium",
        "결과": "",
        "비고": "예외",
        "설명": "",
    },
    {
        "No": 11,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입\n3. 홈택스 인증 완료한 경우\n4. 화면 이탈한 상태에서 스크래핑 완료",
        "title": "카카오 알림톡 동작 확인",
        "steps_actions": "알림톡 상세 화면 버튼 선택 시 동작 확인",
        "steps_result": "예상 환급액 확인 페이지로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 12,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입\n3. 홈택스 인증 완료한 경우\n4. 스크래핑 완료된 상태",
        "title": "예상 환급액 동작 확인",
        "steps_actions": "1. 예상 환급액 0원인 경우 동작 확인\n2. 예상 환급액 1원이상인 경우 동작 확인",
        "steps_result": "1. 예상 환급액 0원 노출되며, 관련 툴팁(0원인 이유, 나중에 환급액이 생긴다면) 노출됨\n2. 예상 환급액 1원 이상 노출되며, 관련 툴팁(예상 환급액을 계산하는 기준) 노출됨",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    {
        "No": 13,
        "preconditions": "1. 카카오 계정 보유\n2. 토스 외 경로로 진입",
        "title": "진행에 어려움이 있으신가요? 영역 동작 확인",
        "steps_actions": "1. 진행에 어려움이 있으신가요? 영역 선택 시 동작 확인",
        "steps_result": "1. 관련 페이지로 이동",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "",
    },
    # 추가 테스트 케이스
    {
        "No": 14,
        "preconditions": "1. 브라우저 실행",
        "title": "홈페이지 로드 확인",
        "steps_actions": "1. https://qa.hiddenmoney.co.kr 접속",
        "steps_result": "페이지가 정상적으로 로드됨",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "자동화: test_home.py",
    },
    {
        "No": 15,
        "preconditions": "1. 홈페이지 접속",
        "title": "페이지 타이틀 확인",
        "steps_actions": "1. 페이지 타이틀 확인",
        "steps_result": "타이틀에 'SAVETAX' 또는 서비스명 포함",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "자동화: test_home.py",
    },
    {
        "No": 16,
        "preconditions": "1. 홈페이지 접속",
        "title": "네비게이션 메뉴 표시 확인",
        "steps_actions": "1. 헤더 영역 확인\n2. 네비게이션 메뉴 확인",
        "steps_result": "네비게이션 메뉴가 정상적으로 표시됨",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "자동화: test_navigation.py",
    },
    {
        "No": 17,
        "preconditions": "1. 홈페이지 접속",
        "title": "푸터 영역 표시 확인",
        "steps_actions": "1. 페이지 하단으로 스크롤\n2. 푸터 영역 확인",
        "steps_result": "푸터가 정상적으로 표시됨",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "자동화: test_navigation.py",
    },
    {
        "No": 18,
        "preconditions": "1. 브라우저 실행",
        "title": "모바일 화면 레이아웃 확인",
        "steps_actions": "1. 화면 크기 375x667 설정\n2. 페이지 로드 확인",
        "steps_result": "모바일 레이아웃 정상 표시",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "자동화: test_ui_elements.py",
    },
    {
        "No": 19,
        "preconditions": "1. 브라우저 실행",
        "title": "태블릿 화면 레이아웃 확인",
        "steps_actions": "1. 화면 크기 768x1024 설정\n2. 페이지 로드 확인",
        "steps_result": "태블릿 레이아웃 정상 표시",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "자동화: test_ui_elements.py",
    },
    {
        "No": 20,
        "preconditions": "1. 브라우저 실행",
        "title": "데스크톱 화면 레이아웃 확인",
        "steps_actions": "1. 화면 크기 1280x800 설정\n2. 페이지 로드 확인",
        "steps_result": "데스크톱 레이아웃 정상 표시",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "자동화: test_ui_elements.py",
    },
    {
        "No": 21,
        "preconditions": "1. 홈페이지 접속",
        "title": "버튼 클릭 가능 여부 확인",
        "steps_actions": "1. 페이지 내 모든 버튼 확인\n2. 버튼 활성화 상태 확인",
        "steps_result": "버튼이 클릭 가능한 상태로 표시됨",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "자동화: test_ui_elements.py",
    },
    {
        "No": 22,
        "preconditions": "1. 홈페이지 접속",
        "title": "이미지 alt 속성 확인 (접근성)",
        "steps_actions": "1. 페이지 내 모든 이미지 확인\n2. alt 속성 존재 여부 확인",
        "steps_result": "이미지에 alt 속성이 존재함",
        "priority": "Low",
        "결과": "",
        "비고": "",
        "설명": "자동화: test_ui_elements.py",
    },
    # ==================== 모바일 UI 테스트 결과 ====================
    {
        "No": 23,
        "preconditions": "1. 브라우저 실행\n2. 화면 크기: 320x568 (mobile_small)",
        "title": "[Mobile] 소형 모바일 UI 분석",
        "steps_actions": "1. 뷰포트 크기 확인\n2. 가로 스크롤 확인\n3. 터치 타겟 크기 확인\n4. 폰트 크기 확인",
        "steps_result": "1. 뷰포트: 500x422\n2. 가로 스크롤 없음\n3. 버튼 3개, 링크 0개\n4. 폰트 크기 적정",
        "priority": "High",
        "결과": "PASS",
        "비고": "터치 타겟 1개 44px 미만",
        "설명": "자동화: test_mobile_ui.py (2026-01-30 실행)",
    },
    {
        "No": 24,
        "preconditions": "1. 브라우저 실행\n2. 화면 크기: 375x667 (mobile)",
        "title": "[Mobile] 일반 모바일 UI 분석",
        "steps_actions": "1. 뷰포트 크기 확인\n2. 가로 스크롤 확인\n3. 터치 타겟 크기 확인\n4. 폰트 크기 확인",
        "steps_result": "1. 뷰포트: 500x521\n2. 가로 스크롤 없음\n3. 버튼 3개, 링크 0개\n4. 폰트 크기 적정",
        "priority": "High",
        "결과": "PASS",
        "비고": "터치 타겟 1개 44px 미만",
        "설명": "자동화: test_mobile_ui.py (2026-01-30 실행)",
    },
    {
        "No": 25,
        "preconditions": "1. 브라우저 실행\n2. 화면 크기: 414x896 (mobile_large)",
        "title": "[Mobile] 대형 모바일 UI 분석",
        "steps_actions": "1. 뷰포트 크기 확인\n2. 가로 스크롤 확인\n3. 터치 타겟 크기 확인\n4. 폰트 크기 확인",
        "steps_result": "1. 뷰포트: 500x671\n2. 가로 스크롤 없음\n3. 버튼 3개, 링크 0개\n4. 폰트 크기 적정",
        "priority": "High",
        "결과": "PASS",
        "비고": "터치 타겟 1개 44px 미만",
        "설명": "자동화: test_mobile_ui.py (2026-01-30 실행)",
    },
    {
        "No": 26,
        "preconditions": "1. 브라우저 실행\n2. 화면 크기: 768x1024 (tablet)",
        "title": "[Mobile] 태블릿 UI 분석",
        "steps_actions": "1. 뷰포트 크기 확인\n2. 가로 스크롤 확인\n3. 터치 타겟 크기 확인\n4. 폰트 크기 확인",
        "steps_result": "1. 뷰포트: 754x671\n2. 가로 스크롤 없음\n3. 버튼 3개, 링크 0개\n4. 폰트 크기 적정",
        "priority": "High",
        "결과": "PASS",
        "비고": "",
        "설명": "자동화: test_mobile_ui.py (2026-01-30 실행)",
    },
    {
        "No": 27,
        "preconditions": "1. 브라우저 실행\n2. 화면 크기: 1280x800 (desktop)",
        "title": "[Mobile] 데스크톱 UI 분석",
        "steps_actions": "1. 뷰포트 크기 확인\n2. 가로 스크롤 확인\n3. 터치 타겟 크기 확인\n4. 폰트 크기 확인",
        "steps_result": "1. 뷰포트: 1266x654\n2. 가로 스크롤 없음\n3. 버튼 3개, 링크 0개\n4. 폰트 크기 적정",
        "priority": "High",
        "결과": "PASS",
        "비고": "",
        "설명": "자동화: test_mobile_ui.py (2026-01-30 실행)",
    },
    {
        "No": 28,
        "preconditions": "1. 브라우저 실행\n2. 화면 크기: 1920x1080 (desktop_large)",
        "title": "[Mobile] 대형 데스크톱 UI 분석",
        "steps_actions": "1. 뷰포트 크기 확인\n2. 가로 스크롤 확인\n3. 터치 타겟 크기 확인\n4. 폰트 크기 확인",
        "steps_result": "1. 뷰포트: 1283x671\n2. 가로 스크롤 없음\n3. 버튼 3개, 링크 0개\n4. 폰트 크기 적정",
        "priority": "High",
        "결과": "PASS",
        "비고": "",
        "설명": "자동화: test_mobile_ui.py (2026-01-30 실행)",
    },
    {
        "No": 29,
        "preconditions": "1. 브라우저 실행\n2. 화면 크기: 375x667 (mobile)",
        "title": "[Mobile] 모바일 메뉴 토글 테스트",
        "steps_actions": "1. 햄버거 메뉴 버튼 검색\n2. 메뉴 클릭 테스트",
        "steps_result": "햄버거 메뉴 버튼 없음 (다른 방식 사용)",
        "priority": "Medium",
        "결과": "PASS",
        "비고": "카카오 로그인 방식으로 메뉴 불필요",
        "설명": "자동화: test_mobile_ui.py (2026-01-30 실행)",
    },
    {
        "No": 30,
        "preconditions": "1. 브라우저 실행\n2. 화면 크기: 375x667 (mobile)",
        "title": "[Mobile] 스크롤 동작 테스트",
        "steps_actions": "1. 페이지 스크롤 높이 확인\n2. 하단으로 스크롤\n3. 스크롤 위치 확인",
        "steps_result": "스크롤 동작 정상",
        "priority": "Medium",
        "결과": "PASS",
        "비고": "",
        "설명": "자동화: test_mobile_ui.py (2026-01-30 실행)",
    },
    {
        "No": 31,
        "preconditions": "1. 브라우저 실행\n2. 화면 크기: 375x667 (mobile)",
        "title": "[Mobile] 뷰포트 메타 태그 확인",
        "steps_actions": "1. viewport 메타 태그 존재 확인\n2. width=device-width 설정 확인\n3. initial-scale=1 설정 확인",
        "steps_result": "width=device-width, initial-scale=1, maximum-scale=1, user-scalable=no",
        "priority": "High",
        "결과": "PASS",
        "비고": "",
        "설명": "자동화: test_mobile_ui.py (2026-01-30 실행)",
    },
    {
        "No": 32,
        "preconditions": "1. 브라우저 실행\n2. 화면 크기: 375x667 (mobile)",
        "title": "[Mobile] 탭 하이라이트 설정 확인",
        "steps_actions": "1. -webkit-tap-highlight-color 속성 확인",
        "steps_result": "tap-highlight-color: rgba(0, 0, 0, 0)",
        "priority": "Low",
        "결과": "PASS",
        "비고": "",
        "설명": "자동화: test_mobile_ui.py (2026-01-30 실행)",
    },
]


def create_excel():
    """TC 케이스 엑셀 파일 생성"""
    # DataFrame 생성
    df = pd.DataFrame(test_cases)

    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "QA 자동화 테스트"

    # 스타일 정의
    header_font = Font(bold=True, color="000000", size=10)
    header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 우선순위 색상
    priority_colors = {
        "High": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    }

    # 헤더 작성
    headers = ["No", "preconditions", "title", "steps_actions", "steps_result", "priority", "결과", "비고", "설명"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 데이터 작성
    for row_idx, tc in enumerate(test_cases, 2):
        for col_idx, header in enumerate(headers, 1):
            value = tc.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # No 컬럼 중앙 정렬
            if header == "No":
                cell.alignment = center_alignment

            # 우선순위 색상
            if header == "priority" and value in priority_colors:
                cell.fill = priority_colors[value]
                cell.alignment = center_alignment

            # 결과/비고 컬럼 중앙 정렬
            if header in ["결과", "비고"]:
                cell.alignment = center_alignment
                if value == "예외":
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 열 너비 설정
    column_widths = {
        "A": 5,    # No
        "B": 25,   # preconditions
        "C": 35,   # title
        "D": 50,   # steps_actions
        "E": 55,   # steps_result
        "F": 10,   # priority
        "G": 8,    # 결과
        "H": 8,    # 비고
        "I": 25,   # 설명
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 행 높이 설정
    ws.row_dimensions[1].height = 25
    for row_idx in range(2, len(test_cases) + 2):
        ws.row_dimensions[row_idx].height = 60

    # 필터 추가
    ws.auto_filter.ref = ws.dimensions

    # 틀 고정
    ws.freeze_panes = "A2"

    # 요약 시트 추가
    ws_summary = wb.create_sheet(title="요약")

    high_count = len([tc for tc in test_cases if tc["priority"] == "High"])
    medium_count = len([tc for tc in test_cases if tc["priority"] == "Medium"])
    low_count = len([tc for tc in test_cases if tc["priority"] == "Low"])

    summary_data = [
        ["SAVETAX QA 테스트 케이스 요약", ""],
        ["", ""],
        ["총 TC 수", len(test_cases)],
        ["", ""],
        ["우선순위별", ""],
        ["High", high_count],
        ["Medium", medium_count],
        ["Low", low_count],
        ["", ""],
        ["생성일", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]

    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif col_idx == 1 and value:
                cell.font = Font(bold=True)

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15

    # 파일 저장
    filepath = "qa/SAVETAX_TC_Cases.xlsx"
    wb.save(filepath)
    print(f"[완료] TC 케이스 엑셀 파일 생성: {filepath}")
    print(f"  - 총 TC 수: {len(test_cases)}")
    print(f"  - High: {high_count}")
    print(f"  - Medium: {medium_count}")
    print(f"  - Low: {low_count}")

    return filepath


if __name__ == "__main__":
    create_excel()
