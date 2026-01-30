"""리포트 생성 모듈

1. Slack 알림
2. 엑셀 리포트 생성
3. HTML 대시보드 (옵션)
"""

import os
import json
import requests
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, asdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class QAReport:
    """QA 리포트 데이터"""
    title: str
    environment: str
    total_tc: int
    passed: int
    failed: int
    skipped: int
    errors: int
    pass_rate: float
    duration_seconds: float
    visual_match_rate: float = 0.0
    api_success_rate: float = 0.0
    performance_score: float = 0.0
    test_results: List[Dict] = None
    visual_diffs: List[Dict] = None
    figma_changes: Dict = None
    generated_at: str = None
    screenshots: List[str] = None

    def __post_init__(self):
        if self.test_results is None:
            self.test_results = []
        if self.visual_diffs is None:
            self.visual_diffs = []
        if self.screenshots is None:
            self.screenshots = []
        if self.generated_at is None:
            self.generated_at = datetime.now().isoformat()


class SlackReporter:
    """Slack 알림 리포터"""

    def __init__(self, webhook_url: str = None):
        self.webhook_url = webhook_url or os.getenv("SLACK_WEBHOOK_URL")

    def send_report(self, report: QAReport) -> bool:
        """Slack으로 리포트 전송"""
        if not self.webhook_url:
            print("SLACK_WEBHOOK_URL이 설정되지 않았습니다.")
            return False

        # 상태에 따른 색상 및 이모지
        if report.pass_rate >= 95:
            color = "good"  # 초록
            emoji = "✅"
            status = "PASSED"
        elif report.pass_rate >= 80:
            color = "warning"  # 노랑
            emoji = "⚠️"
            status = "WARNING"
        else:
            color = "danger"  # 빨강
            emoji = "❌"
            status = "FAILED"

        # 메시지 구성
        message = {
            "attachments": [
                {
                    "color": color,
                    "blocks": [
                        {
                            "type": "header",
                            "text": {
                                "type": "plain_text",
                                "text": f"{emoji} QA 자동화 리포트: {report.title}",
                                "emoji": True
                            }
                        },
                        {
                            "type": "section",
                            "fields": [
                                {"type": "mrkdwn", "text": f"*환경:*\n{report.environment}"},
                                {"type": "mrkdwn", "text": f"*상태:*\n{status}"},
                                {"type": "mrkdwn", "text": f"*통과율:*\n{report.pass_rate:.1f}%"},
                                {"type": "mrkdwn", "text": f"*소요시간:*\n{report.duration_seconds:.1f}초"}
                            ]
                        },
                        {
                            "type": "section",
                            "text": {
                                "type": "mrkdwn",
                                "text": f"*테스트 결과:*\n"
                                        f"• 전체: {report.total_tc}개\n"
                                        f"• 통과: {report.passed}개 ✓\n"
                                        f"• 실패: {report.failed}개 ✗\n"
                                        f"• 에러: {report.errors}개\n"
                                        f"• 스킵: {report.skipped}개"
                            }
                        }
                    ]
                }
            ]
        }

        # 시각적 회귀 테스트 결과 추가
        if report.visual_match_rate > 0:
            message["attachments"][0]["blocks"].append({
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"*시각적 회귀 테스트:*\n"
                            f"• 일치율: {report.visual_match_rate:.1f}%"
                }
            })

        # Figma 변경 사항 추가
        if report.figma_changes and report.figma_changes.get("has_changes"):
            changes = report.figma_changes
            message["attachments"][0]["blocks"].append({
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"*🎨 Figma 변경 감지:*\n"
                            f"• 새 화면: {len(changes.get('new_screens', []))}개\n"
                            f"• 수정된 화면: {len(changes.get('modified_screens', []))}개\n"
                            f"• 삭제된 화면: {len(changes.get('removed_screens', []))}개"
                }
            })

        # 실패한 테스트 목록 (최대 5개)
        failed_tests = [t for t in report.test_results if t.get("status") == "FAIL"][:5]
        if failed_tests:
            failed_text = "*실패한 테스트:*\n"
            for t in failed_tests:
                failed_text += f"• {t.get('title', 'Unknown')}\n"
            message["attachments"][0]["blocks"].append({
                "type": "section",
                "text": {"type": "mrkdwn", "text": failed_text}
            })

        # 타임스탬프
        message["attachments"][0]["blocks"].append({
            "type": "context",
            "elements": [
                {"type": "mrkdwn", "text": f"생성 시간: {report.generated_at}"}
            ]
        })

        # 전송
        try:
            response = requests.post(
                self.webhook_url,
                json=message,
                headers={"Content-Type": "application/json"}
            )
            return response.status_code == 200
        except Exception as e:
            print(f"Slack 전송 실패: {e}")
            return False


class ExcelReporter:
    """엑셀 리포트 생성기"""

    def __init__(self, output_dir: str = None):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl이 필요합니다. `pip install openpyxl` 실행")

        self.output_dir = Path(output_dir) if output_dir else Path(__file__).parent.parent / "reports"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_report(self, report: QAReport) -> Path:
        """엑셀 리포트 생성"""
        wb = Workbook()

        # 1. 요약 시트
        self._create_summary_sheet(wb, report)

        # 2. 테스트 결과 시트
        self._create_results_sheet(wb, report)

        # 3. 시각적 회귀 시트 (있는 경우)
        if report.visual_diffs:
            self._create_visual_sheet(wb, report)

        # 4. Figma 변경 시트 (있는 경우)
        if report.figma_changes and report.figma_changes.get("has_changes"):
            self._create_figma_sheet(wb, report)

        # 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"QA_Report_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)

        print(f"엑셀 리포트 생성: {filepath}")
        return filepath

    def _create_summary_sheet(self, wb: Workbook, report: QAReport):
        """요약 시트 생성"""
        ws = wb.active
        ws.title = "요약"

        # 스타일
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

        # 제목
        ws["A1"] = f"QA 자동화 리포트: {report.title}"
        ws["A1"].font = title_font
        ws.merge_cells("A1:D1")

        # 기본 정보
        info_data = [
            ["환경", report.environment],
            ["생성 시간", report.generated_at],
            ["소요 시간", f"{report.duration_seconds:.1f}초"],
            ["", ""],
            ["테스트 결과", ""],
            ["전체 TC", report.total_tc],
            ["통과", report.passed],
            ["실패", report.failed],
            ["에러", report.errors],
            ["스킵", report.skipped],
            ["통과율", f"{report.pass_rate:.1f}%"],
        ]

        if report.visual_match_rate > 0:
            info_data.append(["", ""])
            info_data.append(["시각적 회귀 테스트", ""])
            info_data.append(["일치율", f"{report.visual_match_rate:.1f}%"])

        for row_idx, (label, value) in enumerate(info_data, 3):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)
            if label and not str(value):
                ws.cell(row=row_idx, column=1).font = header_font

        # 통과율에 따른 색상
        pass_rate_row = 13
        if report.pass_rate >= 95:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif report.pass_rate >= 80:
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        ws.cell(row=pass_rate_row, column=2).fill = fill

        # 열 너비
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

    def _create_results_sheet(self, wb: Workbook, report: QAReport):
        """테스트 결과 시트 생성"""
        ws = wb.create_sheet(title="테스트 결과")

        # 스타일
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        status_colors = {
            "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "FAIL": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "ERROR": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "SKIP": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        }

        # 헤더
        headers = ["No", "테스트 제목", "상태", "소요시간(ms)", "에러 메시지"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # 데이터
        for row_idx, result in enumerate(report.test_results, 2):
            values = [
                result.get("tc_no", row_idx - 1),
                result.get("title", ""),
                result.get("status", ""),
                result.get("duration_ms", 0),
                result.get("error_message", "")[:200]  # 에러 메시지 길이 제한
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # 상태 컬럼 색상
                if col_idx == 3 and value in status_colors:
                    cell.fill = status_colors[value]

        # 열 너비
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 50

        # 필터 추가
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    def _create_visual_sheet(self, wb: Workbook, report: QAReport):
        """시각적 회귀 테스트 시트"""
        ws = wb.create_sheet(title="시각적 회귀")

        headers = ["화면명", "일치 여부", "차이율(%)", "베이스라인", "실제", "Diff"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)

        for row_idx, diff in enumerate(report.visual_diffs, 2):
            ws.cell(row=row_idx, column=1, value=diff.get("screen_name", ""))
            ws.cell(row=row_idx, column=2, value="일치" if diff.get("is_match") else "불일치")
            ws.cell(row=row_idx, column=3, value=diff.get("diff_percentage", 0))
            ws.cell(row=row_idx, column=4, value=diff.get("baseline_path", ""))
            ws.cell(row=row_idx, column=5, value=diff.get("actual_path", ""))
            ws.cell(row=row_idx, column=6, value=diff.get("diff_path", ""))

    def _create_figma_sheet(self, wb: Workbook, report: QAReport):
        """Figma 변경 시트"""
        ws = wb.create_sheet(title="Figma 변경")

        changes = report.figma_changes

        ws["A1"] = "Figma 디자인 변경 사항"
        ws["A1"].font = Font(bold=True, size=12)

        data = [
            ["", ""],
            ["현재 버전", changes.get("current_version", "")],
            ["마지막 수정", changes.get("current_modified", "")],
            ["", ""],
            ["새로 추가된 화면", ", ".join(changes.get("new_screens", []))],
            ["수정된 화면", ", ".join(changes.get("modified_screens", []))],
            ["삭제된 화면", ", ".join(changes.get("removed_screens", []))],
        ]

        for row_idx, (label, value) in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 60


class Reporter:
    """통합 리포터"""

    def __init__(
        self,
        slack_webhook: str = None,
        output_dir: str = None
    ):
        self.slack = SlackReporter(slack_webhook)
        self.excel = ExcelReporter(output_dir) if OPENPYXL_AVAILABLE else None

    def generate_report(
        self,
        title: str,
        environment: str,
        test_result,  # TestSuiteResult
        visual_result=None,  # VisualTestResult
        figma_changes: Dict = None
    ) -> QAReport:
        """통합 리포트 생성"""
        report = QAReport(
            title=title,
            environment=environment,
            total_tc=test_result.total,
            passed=test_result.passed,
            failed=test_result.failed,
            skipped=test_result.skipped,
            errors=test_result.errors,
            pass_rate=test_result.pass_rate,
            duration_seconds=test_result.duration_ms / 1000,
            test_results=[asdict(r) for r in test_result.results],
            figma_changes=figma_changes
        )

        if visual_result:
            report.visual_match_rate = visual_result.match_rate
            report.visual_diffs = [asdict(d) for d in visual_result.diffs]

        return report

    def send_slack(self, report: QAReport) -> bool:
        """Slack 전송"""
        return self.slack.send_report(report)

    def generate_excel(self, report: QAReport) -> Optional[Path]:
        """엑셀 리포트 생성"""
        if self.excel:
            return self.excel.generate_report(report)
        return None

    def report_all(self, report: QAReport) -> Dict:
        """모든 채널로 리포트"""
        results = {
            "slack": self.send_slack(report),
            "excel": None
        }

        excel_path = self.generate_excel(report)
        if excel_path:
            results["excel"] = str(excel_path)

        return results


# CLI 실행
if __name__ == "__main__":
    # 테스트용 더미 데이터
    from qa.automation.test_runner import TestSuiteResult, TestResult

    test_result = TestSuiteResult(
        suite_name="테스트",
        total=10,
        passed=8,
        failed=1,
        skipped=0,
        errors=1,
        duration_ms=5000,
        results=[
            TestResult(tc_no=1, title="테스트 1", status="PASS", duration_ms=100),
            TestResult(tc_no=2, title="테스트 2", status="FAIL", duration_ms=200, error_message="오류 발생"),
        ]
    )

    reporter = Reporter()
    report = reporter.generate_report(
        title="세이브택스 통합 플로우",
        environment="QA",
        test_result=test_result
    )

    # 엑셀 리포트 생성
    excel_path = reporter.generate_excel(report)
    print(f"엑셀 리포트: {excel_path}")

    # Slack 전송 (webhook URL 설정 필요)
    # reporter.send_slack(report)
