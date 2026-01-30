"""통합 전체 플로우_ver.0.3 TC 케이스 엑셀 파일 생성 스크립트"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# TC 케이스 데이터 정의 (Figma 통합 전체 플로우_ver.0.3 기반)
test_cases = [
    # ==================== 1. 통합 홈 화면 ====================
    {
        "No": 1,
        "preconditions": "1. 카카오 로그인 완료\n2. 통합 홈 화면 진입",
        "title": "[통합 홈] 화면 정상 로드 확인",
        "steps_actions": "1. 통합 홈 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. 네비게이션 바 정상 노출\n2. 탭 메뉴 정상 노출\n3. 컨텐츠 영역(섹션1, 섹션2, 안내사항) 정상 노출\n4. 하단 CTA 버튼 정상 노출",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 통합 홈",
    },
    {
        "No": 2,
        "preconditions": "1. 통합 홈 화면 진입",
        "title": "[통합 홈] 탭 네비게이션 동작 확인",
        "steps_actions": "1. 각 탭 메뉴 선택 시 동작 확인\n2. 탭 전환 시 컨텐츠 변경 확인",
        "steps_result": "1. 탭 선택 시 해당 탭 활성화 표시\n2. 탭별 컨텐츠가 정상적으로 전환됨",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 통합 홈 > tab",
    },
    {
        "No": 3,
        "preconditions": "1. 통합 홈 화면 진입\n2. 스크래핑 진행 중",
        "title": "[통합 홈] 스크래핑 진행 상태 표시 확인",
        "steps_actions": "1. 스크래핑 진행 중 화면 확인\n2. 진행 상태 메시지 확인",
        "steps_result": "1. '스크래핑 진행' 상태 표시\n2. 진행 중 안내 메시지 노출",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 통합 홈 > 스크래핑 진행",
    },
    {
        "No": 4,
        "preconditions": "1. 통합 홈 화면 진입",
        "title": "[통합 홈] 환급 진행 현황 영역 확인",
        "steps_actions": "1. 환급 진행 현황 섹션 확인\n2. 예상 환급 금액 표시 확인",
        "steps_result": "1. '환급 진행 현황' 타이틀 노출\n2. 예상 환급액 금액 정상 표시\n3. '예상 환급 금액' 라벨 노출",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 섹션1 > 환급 진행 현황",
    },
    {
        "No": 5,
        "preconditions": "1. 통합 홈 화면 진입",
        "title": "[통합 홈] 헬프 배너 영역 동작 확인",
        "steps_actions": "1. 헬프 배너 영역 확인\n2. '진행에 어려움이 있으신가요?' 버튼 선택",
        "steps_result": "1. 헬프 배너 정상 노출\n2. 버튼 선택 시 관련 도움말 페이지로 이동",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "Figma: 헬프배너",
    },

    # ==================== 2. 브릿지 타입2 화면 ====================
    {
        "No": 6,
        "preconditions": "1. 카카오 로그인 완료\n2. 환급 조회 시작",
        "title": "[브릿지] 환급금 안내 화면 정상 로드 확인",
        "steps_actions": "1. 브릿지 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. '놓쳤던 환급금, 한 번에 찾아드릴게요!' 타이틀 노출\n2. 그래픽 영역 정상 표시\n3. 아코디언 UI 정상 표시",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 브릿지 타입2",
    },
    {
        "No": 7,
        "preconditions": "1. 브릿지 화면 진입",
        "title": "[브릿지] 예상 환급액 확인 버튼 동작 확인",
        "steps_actions": "1. '예상 환급액 확인' 버튼 선택",
        "steps_result": "1. 다음 단계(인증 또는 환급 신청) 화면으로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 브릿지 타입2 > CTA",
    },
    {
        "No": 8,
        "preconditions": "1. 브릿지 화면 진입",
        "title": "[브릿지] 아코디언 UI 동작 확인",
        "steps_actions": "1. 아코디언 영역 선택\n2. 펼침/접힘 동작 확인",
        "steps_result": "1. 아코디언 선택 시 상세 내용 펼쳐짐\n2. 다시 선택 시 접힘",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "Figma: accordion_Type2",
    },

    # ==================== 3. 카카오 가입/로그인 ====================
    {
        "No": 9,
        "preconditions": "1. 서비스 최초 진입\n2. 비로그인 상태",
        "title": "[카카오 가입] 화면 정상 로드 확인",
        "steps_actions": "1. 카카오 가입 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. '세이브택스 환급' 타이틀 노출\n2. '카카오로 계속하기' 버튼 노출\n3. '혹시 법인사업자이신가요?' 링크 노출",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 카카오 가입",
    },
    {
        "No": 10,
        "preconditions": "1. 카카오 가입 화면 진입\n2. 카카오 계정 보유",
        "title": "[카카오 가입] 카카오로 계속하기 버튼 동작 확인",
        "steps_actions": "1. '카카오로 계속하기' 버튼 선택\n2. 카카오 인증 진행",
        "steps_result": "1. 카카오 로그인 화면으로 이동\n2. 로그인 완료 시 서비스 화면으로 복귀",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 카카오 > 카카오로 계속하기",
    },
    {
        "No": 11,
        "preconditions": "1. 카카오 가입 화면 진입\n2. 법인 사업자인 경우",
        "title": "[카카오 가입] 법인사업자 링크 동작 확인",
        "steps_actions": "1. '혹시 법인사업자이신가요?' 링크 선택",
        "steps_result": "1. 법인 사업자 전용 페이지로 이동",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "Figma: 텍스트 버튼 > 혹시 법인사업자이신가요?",
    },

    # ==================== 4. 인증 정보 입력 ====================
    {
        "No": 12,
        "preconditions": "1. 카카오 로그인 완료\n2. 홈택스 인증 필요 상태",
        "title": "[인증] 인증 정보 입력 화면 정상 로드 확인",
        "steps_actions": "1. 인증 정보 입력 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. '어떤 정보를 불러오면 될까요?' 타이틀 노출\n2. '정확한 환급액을 찾으려면 아래 정보의 수집이 필요해요' 안내 문구 노출\n3. 이름 입력 필드 노출",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 인증 정보 입력",
    },
    {
        "No": 13,
        "preconditions": "1. 인증 정보 입력 화면 진입",
        "title": "[인증] 이름 입력 필드 동작 확인",
        "steps_actions": "1. 이름 입력 필드 선택\n2. 이름 입력\n3. 확인 버튼 선택",
        "steps_result": "1. 키보드 노출\n2. 입력값 정상 반영\n3. 다음 단계로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 정보입력1 > 이름을 입력해 주세요",
    },
    {
        "No": 14,
        "preconditions": "1. 인증 정보 입력 화면 진입",
        "title": "[인증] 바텀시트 동작 확인",
        "steps_actions": "1. 바텀시트 트리거 선택\n2. 바텀시트 노출 확인\n3. 바텀시트 닫기 동작 확인",
        "steps_result": "1. 바텀시트 정상 노출\n2. 닫기 버튼 또는 외부 영역 탭 시 바텀시트 닫힘",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "Figma: 인증 정보 입력 > 바텀시트",
    },

    # ==================== 5. 인증 브릿지 ====================
    {
        "No": 15,
        "preconditions": "1. 인증 정보 입력 완료\n2. 인증 진행 중",
        "title": "[인증 브릿지] 인증 진행 화면 정상 로드 확인",
        "steps_actions": "1. 인증 브릿지 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. 인증 진행 중 안내 메시지 노출\n2. 진행 상태 표시",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 인증 브릿지",
    },
    {
        "No": 16,
        "preconditions": "1. 인증 브릿지 화면 진입",
        "title": "[인증 브릿지] 화면 이탈 시 안내 메시지 확인",
        "steps_actions": "1. '화면을 이탈하셔도, 완료가 되면 카카오 알림톡으로 바로 알려드릴게요!' 메시지 확인",
        "steps_result": "1. 안내 메시지 정상 노출",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "Figma: 타이틀 > 화면을 이탈하셔도...",
    },

    # ==================== 6. 환급신청 메인 ====================
    {
        "No": 17,
        "preconditions": "1. 스크래핑 완료\n2. 환급 가능 금액 존재",
        "title": "[환급신청] 화면 정상 로드 확인",
        "steps_actions": "1. 환급신청 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. '환급 신청 시작' 또는 '비사업자 환급' 타이틀 노출\n2. 예상 환급액 표시\n3. CTA 버튼 노출",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청",
    },
    {
        "No": 18,
        "preconditions": "1. 환급신청 화면 진입",
        "title": "[환급신청] 예상 환급액 표시 확인",
        "steps_actions": "1. 예상 환급액 영역 확인\n2. 금액 형식 확인",
        "steps_result": "1. '예상 환급액' 라벨 노출\n2. 금액이 천 단위 콤마로 포맷팅되어 표시 (예: 1,234,567원)",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청 > 예상 환급액 1,234,567원",
    },
    {
        "No": 19,
        "preconditions": "1. 환급신청 화면 진입",
        "title": "[환급신청] 환급 신청 시작하기 버튼 동작 확인",
        "steps_actions": "1. '환급 신청 시작하기' 버튼 선택",
        "steps_result": "1. 계좌 입력 화면으로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: CTA > 환급 신청 시작하기",
    },

    # ==================== 7. 환급신청 계좌입력 ====================
    {
        "No": 20,
        "preconditions": "1. 환급신청 시작\n2. 계좌입력 화면 진입",
        "title": "[계좌입력] 화면 정상 로드 확인",
        "steps_actions": "1. 계좌입력 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. '마지막으로, 환급받을 본인 계좌를 입력해주세요' 타이틀 노출\n2. 은행 선택 필드 노출\n3. 계좌번호 입력 필드 노출\n4. 예상 환급액 표시\n5. '다음' 버튼 노출",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청_계좌입력",
    },
    {
        "No": 21,
        "preconditions": "1. 계좌입력 화면 진입",
        "title": "[계좌입력] 은행 선택 동작 확인",
        "steps_actions": "1. '은행 선택' 필드 선택\n2. 은행 목록 바텀시트 확인\n3. 은행 선택",
        "steps_result": "1. 은행 선택 바텀시트 노출\n2. 은행 목록 정상 표시\n3. 선택한 은행이 필드에 반영",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: icon_bank(input) > 은행 선택",
    },
    {
        "No": 22,
        "preconditions": "1. 계좌입력 화면 진입",
        "title": "[계좌입력] 계좌번호 입력 동작 확인",
        "steps_actions": "1. 계좌번호 입력 필드 선택\n2. 숫자 키보드 노출 확인\n3. 계좌번호 입력",
        "steps_result": "1. 숫자 키보드 노출\n2. 계좌번호 정상 입력\n3. '- 없이 숫자만 입력' 안내 표시",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청_계좌입력(Input_Active_Typing)",
    },
    {
        "No": 23,
        "preconditions": "1. 계좌입력 화면 진입",
        "title": "[계좌입력] 경고 안내 메시지 확인",
        "steps_actions": "1. 하단 안내 메시지 영역 확인",
        "steps_result": "1. '잘못된 계좌, 타인 명의 계좌는 환급이 불가해요' 안내 노출\n2. '세이브택스, 토스뱅크는 신청불가' 안내 노출\n3. '가상계좌번호(휴대폰 번호) 형태의 입력은 안돼요' 안내 노출",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "Figma: icon_dash 영역",
    },
    {
        "No": 24,
        "preconditions": "1. 계좌입력 화면 진입\n2. 은행 미선택 또는 계좌번호 미입력",
        "title": "[계좌입력] 다음 버튼 비활성화 상태 확인",
        "steps_actions": "1. 은행 미선택 상태에서 '다음' 버튼 확인\n2. 계좌번호 미입력 상태에서 '다음' 버튼 확인",
        "steps_result": "1. 은행 미선택 시 '다음' 버튼 비활성화\n2. 계좌번호 미입력 시 '다음' 버튼 비활성화",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: BTN > 다음",
    },
    {
        "No": 25,
        "preconditions": "1. 계좌입력 화면 진입\n2. 은행 선택 완료\n3. 계좌번호 입력 완료",
        "title": "[계좌입력] 다음 버튼 활성화 및 동작 확인",
        "steps_actions": "1. 은행 선택 및 계좌번호 입력\n2. '다음' 버튼 활성화 확인\n3. '다음' 버튼 선택",
        "steps_result": "1. '다음' 버튼 활성화\n2. 접수하기 화면으로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: BTN > 다음",
    },

    # ==================== 8. 환급신청/은행 선택 ====================
    {
        "No": 26,
        "preconditions": "1. 계좌입력 화면에서 은행 선택 필드 탭",
        "title": "[은행 선택] 바텀시트 정상 로드 확인",
        "steps_actions": "1. 은행 선택 바텀시트 확인\n2. 은행 목록 확인",
        "steps_result": "1. 바텀시트 정상 노출\n2. 주요 은행 목록 표시\n3. 스크롤 가능",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청/은행 선택",
    },
    {
        "No": 27,
        "preconditions": "1. 은행 선택 바텀시트 노출",
        "title": "[은행 선택] 은행 선택 동작 확인",
        "steps_actions": "1. 원하는 은행 선택\n2. 바텀시트 닫힘 확인",
        "steps_result": "1. 선택한 은행이 계좌입력 화면에 반영\n2. 바텀시트 자동 닫힘",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청/은행 선택",
    },

    # ==================== 9. 환급신청/접수하기 ====================
    {
        "No": 28,
        "preconditions": "1. 계좌 정보 입력 완료\n2. 접수하기 화면 진입",
        "title": "[접수하기] 화면 정상 로드 확인",
        "steps_actions": "1. 접수하기 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. '아래 내용으로 환급받으시겠어요?' 타이틀 노출\n2. '입력한 정보가 맞는지 확인해주세요' 안내 노출\n3. 예상 환급액 표시\n4. 환급 계좌 정보 표시\n5. '완료하기' 버튼 노출",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청/접수하기",
    },
    {
        "No": 29,
        "preconditions": "1. 접수하기 화면 진입",
        "title": "[접수하기] 입력 정보 확인",
        "steps_actions": "1. 예상 환급액 확인\n2. 환급 계좌 정보(은행명, 계좌번호) 확인",
        "steps_result": "1. 이전에 조회된 예상 환급액과 동일하게 표시\n2. 이전에 입력한 은행명과 계좌번호가 정확히 표시",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청/접수하기 > 예상 환급액, 환급 계좌",
    },
    {
        "No": 30,
        "preconditions": "1. 접수하기 화면 진입",
        "title": "[접수하기] 완료하기 버튼 동작 확인",
        "steps_actions": "1. '완료하기' 버튼 선택",
        "steps_result": "1. 환급 신청 처리\n2. 환급신청완료 화면으로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: CTA > 완료하기",
    },

    # ==================== 10. 환급신청완료 ====================
    {
        "No": 31,
        "preconditions": "1. 환급 신청 완료",
        "title": "[신청완료] 화면 정상 로드 확인",
        "steps_actions": "1. 환급신청완료 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. '마지막으로 환급 신청이 완료되었어요!' 타이틀 노출\n2. '진행상황은 카카오 알림톡으로 알려드릴게요' 안내 노출\n3. 총 예상 환급액 표시\n4. 환급 신청 내역(신청 일시, 계좌 정보 등) 표시\n5. 닫기(X) 버튼 노출",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청완료",
    },
    {
        "No": 32,
        "preconditions": "1. 환급신청완료 화면 진입",
        "title": "[신청완료] 예상 환급액 및 신청 내역 확인",
        "steps_actions": "1. '총 예상 환급액' 확인\n2. 신청 일시 확인\n3. 계좌 정보 확인",
        "steps_result": "1. 총 예상 환급액 금액 표시 (예: 1,184,979원)\n2. '(실제 및 정산내역에 따라 달라질 수 있어요)' 안내 노출\n3. 신청 일시 포맷: 24-05-31-14:00\n4. 환급 계좌 정보 표시",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청완료 > 총 예상 환급액, 신청 일시",
    },
    {
        "No": 33,
        "preconditions": "1. 환급신청완료 화면 진입",
        "title": "[신청완료] 환급 받는 방법 체크박스 확인",
        "steps_actions": "1. '세무서 방문 및 서류 제출' 체크박스 확인",
        "steps_result": "1. 체크박스 정상 노출\n2. 체크/해제 동작 정상",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "Figma: Checkboxes-dark > 세무서 방문 및 서류 제출",
    },
    {
        "No": 34,
        "preconditions": "1. 환급신청완료 화면 진입",
        "title": "[신청완료] 닫기 버튼 동작 확인",
        "steps_actions": "1. 닫기(X) 버튼 선택",
        "steps_result": "1. 통합 홈 화면으로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: touch > icon_close24",
    },
    {
        "No": 35,
        "preconditions": "1. 환급신청완료 화면 진입",
        "title": "[신청완료] 홈으로 이동하기 버튼 동작 확인",
        "steps_actions": "1. '홈으로 이동하기' 버튼 선택",
        "steps_result": "1. 통합 홈 화면으로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: BTN > 홈으로 이동하기",
    },

    # ==================== 11. 환급신청/계좌/완료 ====================
    {
        "No": 36,
        "preconditions": "1. 계좌 등록 완료",
        "title": "[계좌완료] 화면 정상 로드 확인",
        "steps_actions": "1. 계좌 등록 완료 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. 완료 메시지 노출\n2. 총 예상 환급액 표시 (예: 3,100,000원)\n3. 계좌 정보(은행명, 계좌번호) 표시\n4. 신청 일시 표시",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 환급신청/계좌/완료",
    },

    # ==================== 12. 비사업자 종소세 체크리스트 ====================
    {
        "No": 37,
        "preconditions": "1. 비사업자 종소세 환급 대상\n2. 체크리스트 진입",
        "title": "[체크리스트] 구간 체크 브릿지 화면 확인",
        "steps_actions": "1. 구간 체크 브릿지 화면 진입\n2. 화면 구성 요소 확인",
        "steps_result": "1. 브릿지 화면 정상 노출\n2. 안내 메시지 노출\n3. 다음 단계 버튼 노출",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 비사업자 종소세 체크리스트 > 구간 체크 브릿지_1",
    },
    {
        "No": 38,
        "preconditions": "1. 체크리스트 화면 진입",
        "title": "[체크리스트] 질문 템플릿 동작 확인",
        "steps_actions": "1. 질문 내용 확인\n2. '네' 선택 시 동작 확인\n3. '아니요' 선택 시 동작 확인",
        "steps_result": "1. 질문 내용 정상 노출\n2. '네' 선택 시 해당 분기 화면으로 이동\n3. '아니요' 선택 시 해당 분기 화면으로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 질문 템플릿",
    },
    {
        "No": 39,
        "preconditions": "1. 체크리스트 진행 중",
        "title": "[체크리스트] 모든 질문 완료 시 동작 확인",
        "steps_actions": "1. 모든 질문에 응답\n2. 완료 후 화면 전환 확인",
        "steps_result": "1. 중취감 대상인 경우 해당 화면으로 이동\n2. 중취감 대상이 아닌 경우 해당 화면으로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: 모든 질문지/중취감 대상인 경우, 모든 질문지/중취감 대상이 아닌 경우",
    },

    # ==================== 13. 공통 UI/UX ====================
    {
        "No": 40,
        "preconditions": "1. 서비스 내 모든 화면",
        "title": "[공통] 네비게이션 바 뒤로가기 동작 확인",
        "steps_actions": "1. 네비게이션 바 뒤로가기 버튼 선택",
        "steps_result": "1. 이전 화면으로 정상 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: Navigation bar > left",
    },
    {
        "No": 41,
        "preconditions": "1. 서비스 내 모든 화면",
        "title": "[공통] 하단 Chrome(홈 인디케이터) 영역 확인",
        "steps_actions": "1. 하단 Chrome 영역 확인\n2. 홈 인디케이터 표시 확인",
        "steps_result": "1. 하단 Chrome 영역 정상 노출\n2. iOS 홈 인디케이터 정상 표시",
        "priority": "Low",
        "결과": "",
        "비고": "",
        "설명": "Figma: Chrome - Bottom",
    },
    {
        "No": 42,
        "preconditions": "1. 서비스 내 모든 화면",
        "title": "[공통] 상단 Status Bar 영역 확인",
        "steps_actions": "1. 상단 Status Bar 영역 확인\n2. 시간, 배터리 등 표시 확인",
        "steps_result": "1. Status Bar 정상 노출\n2. 시간(9:41 형식), 신호, 배터리 아이콘 정상 표시",
        "priority": "Low",
        "결과": "",
        "비고": "",
        "설명": "Figma: Status Bar - iPhone",
    },

    # ==================== 14. 예외 케이스 ====================
    {
        "No": 43,
        "preconditions": "1. 환급 신청 화면\n2. 인증 요청 메시지 미수신",
        "title": "[예외] 인증 요청 메시지 미수신 시 동작 확인",
        "steps_actions": "1. '인증 요청 메시지를 받지 못하셨나요?' 링크 선택",
        "steps_result": "1. 인증 요청 재시도 또는 대안 안내 화면으로 이동",
        "priority": "Medium",
        "결과": "",
        "비고": "예외",
        "설명": "Figma: 텍스트 버튼 > 인증 요청 메시지를 받지 못하셨나요?",
    },
    {
        "No": 44,
        "preconditions": "1. 신규 유저\n2. 카카오 로그인 완료",
        "title": "[신규가입] 회원가입 플로우 동작 확인",
        "steps_actions": "1. 신규 유저로 서비스 진입\n2. 회원가입 절차 확인",
        "steps_result": "1. '[신규가입] 회원가입' 화면 노출\n2. '[신규가입] 홈택스 인증' 화면으로 이동",
        "priority": "High",
        "결과": "",
        "비고": "",
        "설명": "Figma: [신규가입] 회원가입, [신규가입] 홈택스 인증",
    },
    {
        "No": 45,
        "preconditions": "1. 토스 경로로 진입한 경우",
        "title": "[토스진입] 토스 외 경로로 진입 시 동작 확인",
        "steps_actions": "1. 토스 앱에서 서비스 진입\n2. 레거시 플로우 동작 확인",
        "steps_result": "1. 레거시(기존) 그대로 진행\n2. UI/UX 리뉴얼 적용되지 않음",
        "priority": "Medium",
        "결과": "",
        "비고": "",
        "설명": "Figma: 레거시(기존) 그대로 진행, 토스 외 경로로 진입",
    },
]


def create_excel():
    """TC 케이스 엑셀 파일 생성"""
    # DataFrame 생성
    df = pd.DataFrame(test_cases)

    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "통합 플로우 TC"

    # 스타일 정의
    header_font = Font(bold=True, color="000000", size=10)
    header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 우선순위 색상
    priority_colors = {
        "High": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    }

    # 헤더 작성
    headers = ["No", "preconditions", "title", "steps_actions", "steps_result", "priority", "결과", "비고", "설명"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 데이터 작성
    for row_idx, tc in enumerate(test_cases, 2):
        for col_idx, header in enumerate(headers, 1):
            value = tc.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # No 컬럼 중앙 정렬
            if header == "No":
                cell.alignment = center_alignment

            # 우선순위 색상
            if header == "priority" and value in priority_colors:
                cell.fill = priority_colors[value]
                cell.alignment = center_alignment

            # 결과/비고 컬럼 중앙 정렬
            if header in ["결과", "비고"]:
                cell.alignment = center_alignment
                if value == "예외":
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 열 너비 설정
    column_widths = {
        "A": 5,    # No
        "B": 30,   # preconditions
        "C": 40,   # title
        "D": 50,   # steps_actions
        "E": 55,   # steps_result
        "F": 10,   # priority
        "G": 8,    # 결과
        "H": 8,    # 비고
        "I": 30,   # 설명
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 행 높이 설정
    ws.row_dimensions[1].height = 25
    for row_idx in range(2, len(test_cases) + 2):
        ws.row_dimensions[row_idx].height = 70

    # 필터 추가
    ws.auto_filter.ref = ws.dimensions

    # 틀 고정
    ws.freeze_panes = "A2"

    # 요약 시트 추가
    ws_summary = wb.create_sheet(title="요약")

    high_count = len([tc for tc in test_cases if tc["priority"] == "High"])
    medium_count = len([tc for tc in test_cases if tc["priority"] == "Medium"])
    low_count = len([tc for tc in test_cases if tc["priority"] == "Low"])

    # 카테고리별 TC 수 계산
    categories = {}
    for tc in test_cases:
        title = tc["title"]
        if "[" in title and "]" in title:
            category = title.split("]")[0].replace("[", "")
            categories[category] = categories.get(category, 0) + 1

    summary_data = [
        ["통합 전체 플로우_ver.0.3 TC 요약", ""],
        ["", ""],
        ["총 TC 수", len(test_cases)],
        ["", ""],
        ["우선순위별", ""],
        ["High", high_count],
        ["Medium", medium_count],
        ["Low", low_count],
        ["", ""],
        ["카테고리별", ""],
    ]

    for cat, count in sorted(categories.items()):
        summary_data.append([cat, count])

    summary_data.extend([
        ["", ""],
        ["Figma 원본", "세이브택스 통합 > 통합 전체 플로우_ver.0.3"],
        ["Node ID", "9987-46608"],
        ["", ""],
        ["생성일", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ])

    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif col_idx == 1 and value:
                cell.font = Font(bold=True)

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 40

    # 파일 저장
    filepath = "qa/SAVETAX_UnifiedFlow_TC.xlsx"
    wb.save(filepath)
    print(f"[완료] TC 케이스 엑셀 파일 생성: {filepath}")
    print(f"  - 총 TC 수: {len(test_cases)}")
    print(f"  - High: {high_count}")
    print(f"  - Medium: {medium_count}")
    print(f"  - Low: {low_count}")
    print(f"\n카테고리별:")
    for cat, count in sorted(categories.items()):
        print(f"  - {cat}: {count}")

    return filepath


if __name__ == "__main__":
    create_excel()
